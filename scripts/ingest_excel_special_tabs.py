"""
ingest_excel_special_tabs.py

Indexes the Excel tabs that were skipped during the initial ingest_excel.py run.
Handles a "sectioned" sheet layout where named sub-section labels (e.g., "אודישנים",
"אולפן") appear as single-cell rows interspersed between data rows.

SHEET LAYOUTS
─────────────
Standard  (row 2 = column headers, row 3+ = data)           → tab_type="standard"
Sectioned (row 2 = section label, row 3 = column headers,   → tab_type="sectioned"
           row 4+ = data with more section label rows mixed in)

SKIP LIST  — tabs with no usable data, skipped unconditionally:
    'אור ראשון', 'היורשת', 'המטבח המנצח 2', 'מאסטר שף עונה 9 VIP'

WHAT IT DOES
────────────
1. Adds "section" (searchable) and "tab_type" (filterable) fields to the
   tv-promos index schema (safe, additive — no existing data is changed).
2. Queries the index to find which (show_name, season) pairs are missing.
3. Downloads the Excel, skips already-indexed and SKIP-LIST tabs.
4. Detects sectioned vs standard layout per tab.
5. Parses each missing tab — for sectioned tabs, carries the current section
   label forward and attaches it to every document in that section.
6. Embeds promo_text via Azure OpenAI text-embedding-3-small and uploads.

Usage:
    python ingest_excel_special_tabs.py --preview   # parse + display, no upload
    python ingest_excel_special_tabs.py             # full run
"""

import argparse
import hashlib
import io
import logging
import os
import re
from typing import Any

import httpx
import openpyxl
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchableField,
    SearchFieldDataType,
    SimpleField,
)
from azure.storage.blob import BlobServiceClient
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

AZURE_SEARCH_ENDPOINT         = os.getenv("AZURE_SEARCH_ENDPOINT")
AZURE_SEARCH_KEY              = os.getenv("AZURE_SEARCH_KEY")
AZURE_SEARCH_INDEX_NAME       = os.getenv("AZURE_SEARCH_INDEX_NAME")   # tv-promos
AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
AZURE_STORAGE_CONTAINER_NAME  = os.getenv("AZURE_STORAGE_CONTAINER_NAME")
EXCEL_BLOB_NAME               = os.getenv("EXCEL_BLOB_NAME")

AZURE_OPENAI_ENDPOINT              = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_KEY                   = os.getenv("AZURE_OPENAI_KEY")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT  = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")
AZURE_OPENAI_API_VERSION           = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-01")

EMBED_BATCH_SIZE  = 32
UPLOAD_BATCH_SIZE = 500

# Tabs to skip unconditionally (empty / insufficient data)
SKIP_TABS: frozenset[str] = frozenset({
    "אור ראשון",
    "היורשת",
    "המטבח המנצח 2",
    "מאסטר שף עונה 9 VIP",
})

# Hebrew column header → document field name (must match ingest_excel.py)
COLUMN_MAP: dict[str, str] = {
    "מספר פרק":    "episode_number",
    "מס' פרק":     "episode_number",   # short form used in some tabs
    "יום בשבוע":   "day_of_week",
    "יום":         "day_of_week",       # short form
    "תאריך":       "date",
    "בפרומו":      "promo_text",
    "אודישנים":    "promo_text",        # some tabs use section title as the promo column name
    "נקודת פתיחה": "opening_point",
    "רייטינג פרק":  "rating",
    "תחרות":       "competition",
}
KNOWN_HEADERS: frozenset[str] = frozenset(COLUMN_MAP.keys())

# Strip "עונה N- " prefix from column headers before COLUMN_MAP lookup.
# Handles יצאת צדיק where the column is named "עונה 8- בפרומו" instead of "בפרומו".
_SEASON_PREFIX_RE = re.compile(r"^עונה\s+\d+\s*[-–]\s*")


def _normalize_header(text: str) -> str:
    """Normalise an Excel column header before COLUMN_MAP lookup."""
    return _SEASON_PREFIX_RE.sub("", text).strip()


# Tabs that have NO column header row at all — data starts directly in row 2.
# Column field names are mapped by position (1-based index).
# None = ignore this column position.
POSITIONAL_HEADERS: dict[str, list[str | None]] = {
    "נוטוק": [
        "episode_number", "day_of_week", "date",
        "promo_text", "opening_point", "rating", "competition",
    ],
    "מאסטר שף עונה 11 VIP": [
        "episode_number", "day_of_week", "date",
        "promo_text", "opening_point", "rating", "competition",
    ],
}

# ---------------------------------------------------------------------------
# Tab-name parsing (identical to ingest_excel.py)
# ---------------------------------------------------------------------------

_SEASON_KEYWORD_RE = re.compile(
    r"^(.*?)\s+(?:-\s*)?(?:עונה|עונות)\s+(\d+(?:\s*ו-\s*\d+)?(?:\s+VIP)?)\s*$",
    re.UNICODE,
)
_TRAILING_NUMBER_RE = re.compile(r"^(.+?)\s+(\d+)\s*$", re.UNICODE)


def parse_tab_name(tab_name: str) -> tuple[str, str]:
    tab = tab_name.strip()
    m = _SEASON_KEYWORD_RE.match(tab)
    if m:
        show_name  = m.group(1).strip()
        season_raw = m.group(2).strip()
        season     = re.sub(r"\s*ו-\s*", "-", season_raw)
        return show_name, season
    m = _TRAILING_NUMBER_RE.match(tab)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return tab, ""


# ---------------------------------------------------------------------------
# Helpers (identical to ingest_excel.py)
# ---------------------------------------------------------------------------


def cell_value(cell: Any) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def make_document_id(show_name: str, season: str, episode_number: str, row_index: int) -> str:
    raw = f"{show_name}|{season}|{episode_number}|{row_index}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:40]


def embed_texts(http: httpx.Client, texts: list[str]) -> list[list[float] | None]:
    vectors: list[list[float] | None] = [None] * len(texts)
    non_empty = [(i, t) for i, t in enumerate(texts) if t.strip()]
    if not non_empty:
        return vectors
    url = (
        f"{AZURE_OPENAI_ENDPOINT.rstrip('/')}/openai/deployments/"
        f"{AZURE_OPENAI_EMBEDDING_DEPLOYMENT}/embeddings"
        f"?api-version={AZURE_OPENAI_API_VERSION}"
    )
    for start in range(0, len(non_empty), EMBED_BATCH_SIZE):
        batch = non_empty[start : start + EMBED_BATCH_SIZE]
        indices, batch_texts = zip(*batch)
        resp = http.post(url, json={"input": list(batch_texts)})
        resp.raise_for_status()
        for item in resp.json()["data"]:
            vectors[indices[item["index"]]] = item["embedding"]
    return vectors


# ---------------------------------------------------------------------------
# Tab-type detection
# ---------------------------------------------------------------------------


def _detect_tab_type(sheet: openpyxl.worksheet.worksheet.Worksheet) -> str:
    """
    'standard'  — row 2 col A is a known column header (or empty)
    'sectioned' — row 2 col A is non-empty text that is NOT a column header
    """
    row2_col_a = cell_value(sheet.cell(row=2, column=1))
    if row2_col_a and row2_col_a not in KNOWN_HEADERS:
        return "sectioned"
    return "standard"


def _build_header_map(row) -> dict[int, str]:
    """Build col-index (1-based) → field-name map from a header row."""
    hmap: dict[int, str] = {}
    for col_idx, cell in enumerate(row, start=1):
        text = _normalize_header(cell_value(cell))
        if text in COLUMN_MAP:
            hmap[col_idx] = COLUMN_MAP[text]
    return hmap


def _positional_header_map(tab_name: str) -> dict[int, str]:
    """Return a col-index → field-name map for tabs with no header row."""
    fields = POSITIONAL_HEADERS.get(tab_name, [])
    return {i + 1: f for i, f in enumerate(fields) if f is not None}


def _is_section_label(row, header_map: dict[int, str]) -> tuple[bool, str]:
    """
    Return (True, label_text) when the row is a section-label row.

    A section label row has text only in col A (1-2 non-empty cells total),
    that text is NOT a known column header, and is NOT a bare number
    (episode numbers are numeric; section labels are Hebrew words/phrases).
    """
    non_empty = [(i, cell_value(c)) for i, c in enumerate(row) if cell_value(c)]
    if not non_empty:
        return False, ""            # entirely empty row — not a label

    # More than 2 cells with content → treat as a data row
    if len(non_empty) > 2:
        return False, ""

    first_col_zero_based, first_val = non_empty[0]
    if first_col_zero_based != 0:   # first non-empty not in col A
        return False, ""

    if first_val in KNOWN_HEADERS:  # it's a column header, not a section label
        return False, ""

    # Is it a number (episode number)?
    try:
        float(first_val)
        return False, ""
    except (ValueError, TypeError):
        pass

    return True, first_val


def _is_header_row(row) -> bool:
    """
    Return True when the row is a repeated column-header row (visual separator).

    A header row has ≥ 2 cells whose values exactly match known column header
    names.  Data rows never contain those exact Hebrew header strings as values.
    """
    return sum(1 for c in row if cell_value(c) in KNOWN_HEADERS) >= 2


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------


def _build_doc(show_name: str, season: str, section: str, tab_type: str,
               source_file: str, header_map: dict[int, str],
               row, row_idx: int) -> dict:
    doc: dict[str, str] = {
        "show_name":   show_name,
        "season":      season,
        "section":     section,
        "tab_type":    tab_type,
        "source_file": source_file,
    }
    for col_idx, cell in enumerate(row, start=1):
        field = header_map.get(col_idx)
        if field:
            doc[field] = cell_value(cell)
    for field in COLUMN_MAP.values():
        doc.setdefault(field, "")
    doc["id"] = make_document_id(
        show_name, season, doc.get("episode_number", ""), row_idx
    )
    return doc


def parse_standard_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_file: str,
) -> list[dict]:
    """Parse a standard-layout sheet (col headers in row 2, data from row 3)."""
    show_name, season = parse_tab_name(sheet.title)
    header_map = _build_header_map(sheet[2])
    if not header_map:
        logger.warning(f"    WARNING: no recognised headers in row 2 — skipping '{sheet.title}'.")
        return []

    docs: list[dict] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3), start=3):
        if not any(cell_value(c) for c in row):
            continue
        docs.append(_build_doc(show_name, season, "", "standard",
                               source_file, header_map, row, row_idx))
    return docs


def _find_header_row(sheet) -> tuple[int, dict[int, str]]:
    """
    Scan rows 1-5 to find the first row that contains at least one known column
    header.  Returns (1-based row number, header_map).  Returns (0, {}) if none found.
    """
    for r in range(1, 6):
        hmap = _build_header_map(sheet[r])
        if hmap:
            return r, hmap
    return 0, {}


def _row2_section_info(sheet, header_row: int) -> tuple[str, int]:
    """
    Inspect row 2 to determine the initial section label and first data row.

    Returns (initial_section, data_start):
      - If row 2 col A is a non-numeric, non-header string it is a section label:
          initial_section = that label,  data_start = max(header_row+1, 3)
      - Otherwise row 2 is a data row (e.g. episode 1 when col A = "1"):
          initial_section = "",  data_start = header_row + 1
    """
    row2_col_a = cell_value(sheet.cell(row=2, column=1))
    if row2_col_a and row2_col_a not in KNOWN_HEADERS:
        try:
            float(row2_col_a)
        except (ValueError, TypeError):
            # Non-numeric, non-header text → genuine section label
            return row2_col_a, max(header_row + 1, 3)
    # Numeric or empty → row 2 is a data row; start scanning there
    return "", header_row + 1


def parse_sectioned_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_file: str,
) -> list[dict]:
    """
    Parse a sectioned-layout sheet.

    Four confirmed layouts in the wild:
        A  (headers in row 1, section label in row 2):
               row 1  — column headers (תאריך, בפרומו, …)
               row 2  — first section label  (e.g., "אודישנים")
               row 3+ — data rows interspersed with section-label rows

        B  (headers in row 3, original assumption):
               row 1  — merged show-title header  (skipped)
               row 2  — first section label
               row 3  — column headers
               row 4+ — data rows interspersed with section-label rows

        C  (headers in row 1, first data row in row 2 — no section labels):
               row 1  — column headers
               row 2  — first data row  (col A = episode number like "1")
               row 3+ — more data rows

        D  (NO header row — tabs in POSITIONAL_HEADERS):
               row 1  — merged show-title header  (skipped)
               row 2  — first data row  (episode 1)
               row 3+ — more data rows
               Columns are mapped by position using POSITIONAL_HEADERS.

    _find_header_row() auto-detects the header row.
    _row2_section_info() determines whether row 2 is a section label or a data row.
    """
    show_name, season = parse_tab_name(sheet.title)

    header_row, header_map = _find_header_row(sheet)
    if not header_map:
        # Layout D — no header row; use positional mapping if available
        pos_map = _positional_header_map(sheet.title)
        if not pos_map:
            logger.warning(f"    WARNING: no recognised headers in rows 1-5 — skipping '{sheet.title}'.")
            return []
        logger.info(f"    Using positional column map for '{sheet.title}'.")
        header_map = pos_map
        header_row = 1   # merged title is row 1; data starts at row 2
        current_section = ""
        data_start = 2
    else:
        current_section, data_start = _row2_section_info(sheet, header_row)

    docs: list[dict] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start), start=data_start):
        if not any(cell_value(c) for c in row):
            continue                                # empty row

        if _is_header_row(row):
            continue                                # repeated visual header — skip silently

        is_label, label = _is_section_label(row, header_map)
        if is_label:
            current_section = label
            continue

        docs.append(_build_doc(show_name, season, current_section, "sectioned",
                               source_file, header_map, row, row_idx))
    return docs


# ---------------------------------------------------------------------------
# Preview — print the parsed structure of a tab
# ---------------------------------------------------------------------------


def preview_tab(wb: openpyxl.Workbook, tab_name: str) -> None:
    sheet = wb[tab_name]
    tab_type = _detect_tab_type(sheet)
    show_name, season = parse_tab_name(tab_name)

    logger.info(f"\n{'=' * 64}")
    logger.info(f"  TAB PREVIEW: '{tab_name}'")
    logger.info(f"  Type      : {tab_type}")
    logger.info(f"  show_name : {show_name!r}")
    logger.info(f"  season    : {season!r}")
    logger.info(f"{'=' * 64}")

    if tab_type == "standard":
        header_map = _build_header_map(sheet[2])
        data_rows  = sum(
            1 for row in sheet.iter_rows(min_row=3)
            if any(cell_value(c) for c in row)
        )
        logger.info(f"  Headers (row 2) : {list(header_map.values())}")
        logger.info(f"  Data rows       : {data_rows}")
        return

    # Sectioned
    header_row, header_map = _find_header_row(sheet)
    current_section, data_start = _row2_section_info(sheet, header_row)
    logger.info(f"  Headers (row {header_row}) : {list(header_map.values())}")

    # Walk from data_start collecting section labels and row counts
    sections: dict[str, int] = {}
    sections[current_section] = 0

    for row in sheet.iter_rows(min_row=data_start):
        if not any(cell_value(c) for c in row):
            continue
        is_label, label = _is_section_label(row, header_map)
        if is_label:
            current_section = label
            sections.setdefault(current_section, 0)
        else:
            sections[current_section] = sections.get(current_section, 0) + 1

    logger.info(f"\n  Sections found ({len(sections)}):")
    for sec_name, count in sections.items():
        logger.info(f"    [{sec_name}]  — {count} data row(s)")

    # Sample: first 3 data rows
    logger.info(f"\n  Sample rows (first 3):")
    shown = 0
    # current_section already set by _row2_section_info above
    for row in sheet.iter_rows(min_row=data_start):
        if shown >= 3:
            break
        if not any(cell_value(c) for c in row):
            continue
        is_label, label = _is_section_label(row, header_map)
        if is_label:
            current_section = label
            continue
        values = {
            fname: cell_value(row[col_idx - 1])
            for col_idx, fname in header_map.items()
            if col_idx <= len(row)
        }
        logger.info(f"    section={current_section!r}  {values}")
        shown += 1


# ---------------------------------------------------------------------------
# Row-by-row trace (debugging aid)
# ---------------------------------------------------------------------------


def trace_tab(wb: openpyxl.Workbook, tab_name: str) -> None:
    """
    Print a row-by-row classification table for a sectioned tab.

    Classification per row:
        EMPTY          — all cells empty
        HEADER_REPEAT  — ≥2 cells match known column header names  (skip)
        SECTION_LABEL  — single text label in col A                (update section)
        DATA           — everything else                           (→ document)
    """
    sheet = wb[tab_name]
    header_row_num, header_map = _find_header_row(sheet)
    current_section, data_start = _row2_section_info(sheet, header_row_num)

    COL_W = 28   # truncation width for cell value display

    logger.info(f"\n{'=' * 80}")
    logger.info(f"  TRACE: '{tab_name}'")
    logger.info(f"  headers in row {header_row_num}: {list(header_map.values())}")
    logger.info(f"  first section (row 2): '{current_section}'")
    logger.info(f"  data scan starts at row {data_start}")
    logger.info(f"{'=' * 80}")
    logger.info(f"  {'ROW':>4}  {'NON-EMPTY':>9}  {'CLASSIFICATION':<20}  "
                f"{'COL_A':<{COL_W}}  {'COL_B':<{COL_W}}  COL_C")
    logger.info(f"  {'-'*4}  {'-'*9}  {'-'*20}  {'-'*COL_W}  {'-'*COL_W}  {'-'*20}")

    for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start), start=data_start):
        all_vals = [cell_value(c) for c in row]
        non_empty_count = sum(1 for v in all_vals if v)

        # --- classify ---
        if non_empty_count == 0:
            cls = "EMPTY"
        elif _is_header_row(row):
            cls = "HEADER_REPEAT"
        else:
            is_label, label = _is_section_label(row, header_map)
            if is_label:
                current_section = label
                cls = "SECTION_LABEL"
            else:
                cls = f"DATA"

        # --- display first 3 cell values, truncated ---
        def _trunc(s: str, w: int = COL_W) -> str:
            return (s[:w - 1] + "…") if len(s) > w else s

        c1 = _trunc(all_vals[0]) if len(all_vals) > 0 else ""
        c2 = _trunc(all_vals[1]) if len(all_vals) > 1 else ""
        c3 = _trunc(all_vals[2]) if len(all_vals) > 2 else ""

        section_tag = f" [{current_section[:18]}]" if cls == "DATA" else ""
        logger.info(f"  {row_idx:>4}  {non_empty_count:>9}  "
                    f"{cls + section_tag:<20}  {c1:<{COL_W}}  {c2:<{COL_W}}  {c3}")

    logger.info(f"{'=' * 80}")


# ---------------------------------------------------------------------------
# Index schema update
# ---------------------------------------------------------------------------


def update_index_schema(index_client: SearchIndexClient, index_name: str) -> None:
    """
    Add 'section' and 'tab_type' fields to the index if not already present.
    Uses create_or_update_index which is additive — no existing data is changed.
    """
    index = index_client.get_index(index_name)
    existing = {f.name for f in index.fields}
    new_fields = []

    if "section" not in existing:
        new_fields.append(SearchableField(
            name="section",
            type=SearchFieldDataType.String,
            retrievable=True,
            filterable=True,
            analyzer_name="he.microsoft",
        ))
    if "tab_type" not in existing:
        new_fields.append(SimpleField(
            name="tab_type",
            type=SearchFieldDataType.String,
            retrievable=True,
            filterable=True,
        ))

    if not new_fields:
        logger.info("  Index already has 'section' and 'tab_type' fields — no schema change needed.")
        return

    index.fields = list(index.fields) + new_fields
    index_client.create_or_update_index(index)
    names = [f.name for f in new_fields]
    logger.info(f"  Index schema updated — added fields: {names}")


# ---------------------------------------------------------------------------
# Index query helpers
# ---------------------------------------------------------------------------


def get_indexed_pairs(search_client: SearchClient) -> set[tuple[str, str]]:
    """Return the set of (show_name, season) pairs already in the index."""
    pairs: set[tuple[str, str]] = set()
    for doc in search_client.search("*", select=["show_name", "season"]):
        pairs.add((doc.get("show_name", "") or "", doc.get("season", "") or ""))
    return pairs


# ---------------------------------------------------------------------------
# Upload helpers
# ---------------------------------------------------------------------------


def upload_in_batches(search_client: SearchClient, docs: list[dict]) -> int:
    total = 0
    for start in range(0, len(docs), UPLOAD_BATCH_SIZE):
        batch   = docs[start : start + UPLOAD_BATCH_SIZE]
        results = search_client.upload_documents(documents=batch)
        ok      = sum(1 for r in results if r.succeeded)
        failed  = len(batch) - ok
        total  += ok
        if failed:
            logger.warning(f"    WARNING: {failed} document(s) failed to upload.")
    return total


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main(preview: bool = False, trace: bool = False, retab: list[str] | None = None) -> None:
    required = {
        "AZURE_SEARCH_ENDPOINT":          AZURE_SEARCH_ENDPOINT,
        "AZURE_SEARCH_KEY":               AZURE_SEARCH_KEY,
        "AZURE_SEARCH_INDEX_NAME":        AZURE_SEARCH_INDEX_NAME,
        "AZURE_STORAGE_CONNECTION_STRING": AZURE_STORAGE_CONNECTION_STRING,
        "AZURE_STORAGE_CONTAINER_NAME":   AZURE_STORAGE_CONTAINER_NAME,
        "EXCEL_BLOB_NAME":                EXCEL_BLOB_NAME,
        "AZURE_OPENAI_ENDPOINT":          AZURE_OPENAI_ENDPOINT,
        "AZURE_OPENAI_KEY":               AZURE_OPENAI_KEY,
        "AZURE_OPENAI_EMBEDDING_DEPLOYMENT": AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        raise EnvironmentError(
            f"Missing required environment variables: {', '.join(missing)}. "
            "Check your .env file."
        )

    # ---- Clients ----
    credential    = AzureKeyCredential(AZURE_SEARCH_KEY)
    search_client = SearchClient(AZURE_SEARCH_ENDPOINT, AZURE_SEARCH_INDEX_NAME, credential)
    index_client  = SearchIndexClient(AZURE_SEARCH_ENDPOINT, credential)
    blob_service  = BlobServiceClient.from_connection_string(AZURE_STORAGE_CONNECTION_STRING)
    openai_http   = httpx.Client(
        headers={"api-key": AZURE_OPENAI_KEY, "Content-Type": "application/json"},
        timeout=60.0,
    )

    # ---- Update index schema (additive, safe to run every time) ----
    logger.info("Checking index schema ...")
    update_index_schema(index_client, AZURE_SEARCH_INDEX_NAME)

    # ---- Download Excel ----
    logger.info(f"\nDownloading '{EXCEL_BLOB_NAME}' ...")
    blob_client  = blob_service.get_blob_client(
        container=AZURE_STORAGE_CONTAINER_NAME, blob=EXCEL_BLOB_NAME
    )
    excel_bytes  = blob_client.download_blob().readall()
    wb           = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    logger.info(f"  {len(wb.sheetnames)} sheet(s) in workbook.")

    # ---- Find which tabs are missing from the index ----
    logger.info("\nQuerying index for already-indexed (show_name, season) pairs ...")
    indexed_pairs = get_indexed_pairs(search_client)
    logger.info(f"  {len(indexed_pairs)} pair(s) already indexed.")

    # Tabs explicitly requested for re-indexing (already in index but data needs refresh)
    force_retab: set[str] = set(retab or [])
    if force_retab:
        missing = force_retab - set(wb.sheetnames)
        if missing:
            logger.warning(f"  WARNING: --retab tabs not found in workbook: {missing}")
        logger.info(f"  Force re-index  : {sorted(force_retab)} tab(s)")

    tabs_to_process: list[str] = []
    tabs_skipped_list: list[str] = []
    tabs_already_indexed: list[str] = []

    for tab_name in wb.sheetnames:
        if tab_name in SKIP_TABS and tab_name not in force_retab:
            tabs_skipped_list.append(tab_name)
            continue
        show_name, season = parse_tab_name(tab_name)
        if (show_name, season) in indexed_pairs and tab_name not in force_retab:
            tabs_already_indexed.append(tab_name)
        else:
            tabs_to_process.append(tab_name)

    logger.info(f"\n  Already indexed : {len(tabs_already_indexed)} tab(s) — will skip")
    logger.info(f"  In skip list    : {len(tabs_skipped_list)} tab(s) — will skip")
    logger.info(f"  To process      : {len(tabs_to_process)} tab(s)")
    if tabs_to_process:
        for t in tabs_to_process:
            tab_type = _detect_tab_type(wb[t])
            forced = " [FORCED RE-INDEX]" if t in force_retab else ""
            logger.info(f"    '{t}'  [{tab_type}]{forced}")

    # ---- Trace mode: row-by-row classification for each tab-to-process ----
    if trace:
        for tab_name in tabs_to_process:
            trace_tab(wb, tab_name)
        logger.info("\n-- TRACE MODE: no documents uploaded --")
        return

    # ---- Always preview 'מאסטר שף עונה 10' if it's in the workbook ----
    SPECIAL_PREVIEW_TAB = "מאסטר שף עונה 10"
    if SPECIAL_PREVIEW_TAB in wb.sheetnames:
        preview_tab(wb, SPECIAL_PREVIEW_TAB)

    # Additional previews for any other tab-to-process in preview mode
    if preview:
        for tab_name in tabs_to_process:
            if tab_name != SPECIAL_PREVIEW_TAB:
                preview_tab(wb, tab_name)
        logger.info("\n-- PREVIEW MODE: no documents uploaded --")
        return

    if not tabs_to_process:
        logger.info("\nNothing to do — all non-skipped tabs are already indexed.")
        return

    # ---- Parse, embed, upload ----
    logger.info("")
    grand_total  = 0
    tabs_done    = 0
    tabs_errored = 0

    for tab_name in tabs_to_process:
        sheet    = wb[tab_name]
        tab_type = _detect_tab_type(sheet)
        logger.info(f"  Processing '{tab_name}'  [{tab_type}] ...")

        try:
            if tab_type == "sectioned":
                docs = parse_sectioned_sheet(sheet, EXCEL_BLOB_NAME)
            else:
                docs = parse_standard_sheet(sheet, EXCEL_BLOB_NAME)

            if not docs:
                logger.info(f"    No documents produced — skipping.")
                continue

            # Embed promo_text for all docs in this tab
            promo_texts = [d.get("promo_text", "") for d in docs]
            vectors     = embed_texts(openai_http, promo_texts)
            embedded    = 0
            for doc, vec in zip(docs, vectors):
                if vec is not None:
                    doc["promo_vector"] = vec
                    embedded += 1

            uploaded = upload_in_batches(search_client, docs)
            grand_total += uploaded
            tabs_done   += 1
            logger.info(f"    Parsed {len(docs)} row(s), embedded {embedded}, "
                        f"uploaded {uploaded}.")

        except Exception as exc:
            logger.error(f"    ERROR: {exc}")
            tabs_errored += 1

    logger.info("")
    logger.info("=" * 52)
    logger.info(f"  Tabs processed  : {tabs_done}")
    logger.info(f"  Tabs errored    : {tabs_errored}")
    logger.info(f"  Total docs uploaded : {grand_total}")
    logger.info("=" * 52)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Index special-structure Excel tabs into tv-promos"
    )
    parser.add_argument(
        "--preview",
        action="store_true",
        help="Parse and display structure without uploading to the index",
    )
    parser.add_argument(
        "--trace",
        action="store_true",
        help="Print row-by-row classification for each tab-to-process (no upload)",
    )
    parser.add_argument(
        "--retab",
        metavar="TAB_NAME",
        nargs="+",
        help=(
            "Force re-index one or more already-indexed tab(s). "
            "Existing documents for the tab's (show_name, season) pair are NOT deleted first — "
            "Azure AI Search upserts by document key so updated rows overwrite stale ones. "
            "Example: --retab \"יצאת צדיק עונה 8\" \"יצאת צדיק עונה 9\""
        ),
    )
    args = parser.parse_args()
    main(preview=args.preview, trace=args.trace, retab=args.retab)
