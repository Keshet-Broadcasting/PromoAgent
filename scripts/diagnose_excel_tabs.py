"""
diagnose_excel_tabs.py

Diagnoses which Excel tabs from "מעקבי פרומו.xlsx" were NOT indexed into
the "tv-promos" Azure AI Search index and why.

The index stores (show_name, season) parsed from the tab title — not the raw
tab name — so the comparison reconstructs those values using the same
parse_tab_name() logic used during ingestion.

Output:
  - Total sheets in Excel vs. indexed
  - Missing tabs (tab name + parsed show_name/season + likely cause)
  - Extra entries in the index not matching any current Excel tab
  - Per-missing-tab detail: row count, header row presence

Usage:
    python diagnose_excel_tabs.py
"""

import io
import logging
import os
import re

import openpyxl
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.storage.blob import BlobServiceClient
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

AZURE_SEARCH_ENDPOINT   = os.getenv("AZURE_SEARCH_ENDPOINT")
AZURE_SEARCH_KEY        = os.getenv("AZURE_SEARCH_KEY")
AZURE_SEARCH_INDEX_NAME = os.getenv("AZURE_SEARCH_INDEX_NAME")          # tv-promos
AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
AZURE_STORAGE_CONTAINER_NAME    = os.getenv("AZURE_STORAGE_CONTAINER_NAME")
EXCEL_BLOB_NAME = os.getenv("EXCEL_BLOB_NAME")

# Hebrew column headers recognised by the ingestion script
KNOWN_HEADERS = {
    "מספר פרק", "יום בשבוע", "תאריך", "בפרומו",
    "נקודת פתיחה", "רייטינג פרק", "תחרות",
}


# ---------------------------------------------------------------------------
# parse_tab_name — identical copy from ingest_excel.py
# ---------------------------------------------------------------------------

_SEASON_KEYWORD_RE = re.compile(
    r"^(.*?)\s+(?:-\s*)?(?:עונה|עונות)\s+(\d+(?:\s*ו-\s*\d+)?(?:\s+VIP)?)\s*$",
    re.UNICODE,
)
_TRAILING_NUMBER_RE = re.compile(r"^(.+?)\s+(\d+)\s*$", re.UNICODE)


def parse_tab_name(tab_name: str) -> tuple[str, str]:
    tab = tab_name.strip()
    m = _SEASON_KEYWORD_RE.match(tab)
    if m:
        show_name  = m.group(1).strip()
        season_raw = m.group(2).strip()
        season     = re.sub(r"\s*ו-\s*", "-", season_raw)
        return show_name, season
    m = _TRAILING_NUMBER_RE.match(tab)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return tab, ""


# ---------------------------------------------------------------------------
# Step 1 — download the Excel workbook
# ---------------------------------------------------------------------------


def download_workbook() -> openpyxl.Workbook:
    logger.info(f"Downloading '{EXCEL_BLOB_NAME}' from container '{AZURE_STORAGE_CONTAINER_NAME}' ...")
    blob_service = BlobServiceClient.from_connection_string(AZURE_STORAGE_CONNECTION_STRING)
    blob_client  = blob_service.get_blob_client(
        container=AZURE_STORAGE_CONTAINER_NAME, blob=EXCEL_BLOB_NAME
    )
    data = blob_client.download_blob().readall()
    logger.info(f"  Downloaded {len(data):,} bytes.")
    return openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True, read_only=True)


# ---------------------------------------------------------------------------
# Step 2 — collect all (show_name, season) pairs present in the index
# ---------------------------------------------------------------------------


def fetch_indexed_pairs(search_client: SearchClient) -> set[tuple[str, str]]:
    """
    Iterate ALL documents in the index (auto-paginated by the SDK) and return
    the set of unique (show_name, season) pairs.
    """
    logger.info("Querying 'tv-promos' index for all indexed (show_name, season) pairs ...")
    pairs: set[tuple[str, str]] = set()
    results = search_client.search(
        search_text="*",
        select=["show_name", "season"],
    )
    for doc in results:
        pairs.add((doc.get("show_name", "") or "", doc.get("season", "") or ""))
    logger.info(f"  Found {len(pairs)} unique (show_name, season) pair(s) in the index.")
    return pairs


# ---------------------------------------------------------------------------
# Step 3 — per-tab detail for missing tabs
# ---------------------------------------------------------------------------


def _cell_text(cell) -> str:
    return str(cell.value).strip() if cell.value is not None else ""


def inspect_sheet(wb: openpyxl.Workbook, sheet_name: str) -> dict:
    """
    Return a dict with diagnostic info about the sheet:
        data_rows   — number of non-empty rows from row 3 onward
        headers     — list of non-empty values found in row 2
        known_hdrs  — number of headers matching the expected column names
    """
    sheet = wb[sheet_name]

    # Row 2 — column headers
    headers = [_cell_text(c) for c in sheet[2] if _cell_text(c)]
    known   = sum(1 for h in headers if h in KNOWN_HEADERS)

    # Rows 3+ — count non-empty rows (stop at first 500 to avoid scanning huge sheets)
    data_rows = 0
    for i, row in enumerate(sheet.iter_rows(min_row=3), start=3):
        if any(c.value is not None for c in row):
            data_rows += 1
        if i > 502:
            data_rows = f"{data_rows}+ (stopped counting at row 502)"
            break

    return {
        "data_rows":  data_rows,
        "headers":    headers,
        "known_hdrs": known,
    }


def _likely_cause(show_name: str, season: str, detail: dict) -> str:
    """Infer the most likely reason a tab wasn't indexed."""
    data_rows = detail["data_rows"]
    no_data   = data_rows == 0 or data_rows == "0"
    no_hdrs   = detail["known_hdrs"] == 0

    if not show_name:
        return "parse_tab_name() returned empty show_name"
    if no_data:
        return "no data rows found (empty sheet)"
    if no_hdrs:
        return f"row 2 headers not recognised — found: {detail['headers'][:5]}"
    return "unknown — check ingest_excel.py warnings for this tab"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    missing_vars = [
        v for v in (
            "AZURE_SEARCH_ENDPOINT", "AZURE_SEARCH_KEY", "AZURE_SEARCH_INDEX_NAME",
            "AZURE_STORAGE_CONNECTION_STRING", "AZURE_STORAGE_CONTAINER_NAME",
            "EXCEL_BLOB_NAME",
        )
        if not os.getenv(v)
    ]
    if missing_vars:
        raise EnvironmentError(
            f"Missing required environment variables: {', '.join(missing_vars)}. "
            "Check your .env file."
        )

    # ---- Download Excel ----
    wb = download_workbook()
    excel_sheets = wb.sheetnames
    logger.info(f"  Workbook has {len(excel_sheets)} sheet(s).\n")

    # ---- Build Excel → (show_name, season) map ----
    # key = tab name, value = parsed pair
    excel_map: dict[str, tuple[str, str]] = {
        tab: parse_tab_name(tab) for tab in excel_sheets
    }

    # Detect duplicate parsed pairs (two tabs mapping to the same key)
    seen_pairs: dict[tuple[str, str], list[str]] = {}
    for tab, pair in excel_map.items():
        seen_pairs.setdefault(pair, []).append(tab)
    duplicates = {pair: tabs for pair, tabs in seen_pairs.items() if len(tabs) > 1}

    # ---- Query index ----
    credential    = AzureKeyCredential(AZURE_SEARCH_KEY)
    search_client = SearchClient(
        endpoint=AZURE_SEARCH_ENDPOINT,
        index_name=AZURE_SEARCH_INDEX_NAME,
        credential=credential,
    )
    indexed_pairs = fetch_indexed_pairs(search_client)

    # ---- Compare ----
    excel_pairs = set(excel_map.values())
    missing_pairs = excel_pairs - indexed_pairs
    extra_pairs   = indexed_pairs - excel_pairs

    # Map missing pairs back to their tab names
    pair_to_tabs: dict[tuple[str, str], list[str]] = {}
    for tab, pair in excel_map.items():
        pair_to_tabs.setdefault(pair, []).append(tab)

    missing_tabs: list[tuple[str, tuple[str, str]]] = []
    for pair in missing_pairs:
        for tab in pair_to_tabs.get(pair, [f"(unknown — parsed as {pair})"]):
            missing_tabs.append((tab, pair))
    missing_tabs.sort(key=lambda x: x[0])

    # ---- Print summary ----
    sep = "=" * 60
    logger.info(sep)
    logger.info(f"  Total sheets in Excel : {len(excel_sheets)}")
    logger.info(f"  Unique parsed keys    : {len(excel_pairs)}")
    logger.info(f"  Indexed pairs         : {len(indexed_pairs)}")
    logger.info(f"  Missing from index    : {len(missing_pairs)}")
    logger.info(f"  Extra in index        : {len(extra_pairs)}")
    logger.info(sep)

    # ---- Duplicate parse warnings ----
    if duplicates:
        logger.warning("\nWARNING — multiple tabs parse to the same (show_name, season):")
        for (sn, se), tabs in duplicates.items():
            logger.info(f"  show_name={sn!r}  season={se!r}")
            for t in tabs:
                logger.info(f"    tab: {t!r}")

    # ---- Missing tabs detail ----
    if missing_tabs:
        logger.info(f"\nMISSING tabs ({len(missing_tabs)}) — not found in index:")
        logger.info("-" * 60)
        for tab, (show_name, season) in missing_tabs:
            detail = inspect_sheet(wb, tab)
            cause  = _likely_cause(show_name, season, detail)
            logger.info(f"  Tab       : {tab!r}")
            logger.info(f"  Parsed as : show_name={show_name!r}  season={season!r}")
            logger.info(f"  Data rows : {detail['data_rows']}")
            logger.info(f"  Row-2 hdrs: {detail['headers']}")
            logger.info(f"  Likely cause: {cause}")
            logger.info("")
    else:
        logger.info("\nAll Excel tabs are represented in the index.")

    # ---- Extra in index ----
    if extra_pairs:
        logger.info(f"\nEXTRA in index ({len(extra_pairs)}) — not in current Excel file:")
        logger.info("-" * 60)
        for show_name, season in sorted(extra_pairs):
            logger.info(f"  show_name={show_name!r}  season={season!r}")

    logger.info(sep)


if __name__ == "__main__":
    main()
