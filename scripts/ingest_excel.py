"""
ingest_excel.py

Downloads the Excel promo-tracking file from Azure Blob Storage, parses every
tab, and uploads each data row as a separate document to Azure AI Search.

Supported tab title formats (Hebrew):
    [SHOW] עונה N          →  e.g. "ארץ נהדרת עונה 18"
    [SHOW] עונה N VIP      →  e.g. "מאסטר שף עונה 9 VIP"
    [SHOW] - עונות N ו-M   →  e.g. "המטבח המנצח VIP - עונות 2 ו-3"  (season → "N-M")
    [SHOW] N               →  e.g. "המתמחים 3"  (bare trailing number)
    [SHOW]                 →  e.g. "אור ראשון"  (no season)

Excel column layout (row 1 = merged title, row 2 = column headers):
    מספר פרק | יום בשבוע | תאריך | בפרומו | נקודת פתיחה | רייטינג פרק | תחרות

Usage:
    python ingest_excel.py
"""

import hashlib
import io
import logging
import os
import re
from typing import Any

import openpyxl
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.storage.blob import BlobServiceClient
import httpx
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

AZURE_SEARCH_ENDPOINT = os.getenv("AZURE_SEARCH_ENDPOINT")
AZURE_SEARCH_KEY = os.getenv("AZURE_SEARCH_KEY")
AZURE_SEARCH_INDEX_NAME = os.getenv("AZURE_SEARCH_INDEX_NAME")
AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
AZURE_STORAGE_CONTAINER_NAME = os.getenv("AZURE_STORAGE_CONTAINER_NAME")
EXCEL_BLOB_NAME = os.getenv("EXCEL_BLOB_NAME")

AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_KEY = os.getenv("AZURE_OPENAI_KEY")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-01")

# How many texts to embed in a single Azure OpenAI request (max 2048; keep low to
# stay within per-request token limits and avoid throttling)
EMBED_BATCH_SIZE = 32

# Pattern 1 — explicit season keyword (עונה singular, עונות plural).
# Optional " - " separator before the keyword; optional range "ו-N"; optional " VIP" suffix.
# Examples: "ארץ נהדרת עונה 18", "מאסטר שף עונה 9 VIP", "המטבח המנצח VIP - עונות 2 ו-3"
_SEASON_KEYWORD_RE = re.compile(
    r"^(.*?)\s+(?:-\s*)?(?:עונה|עונות)\s+(\d+(?:\s*ו-\s*\d+)?(?:\s+VIP)?)\s*$",
    re.UNICODE,
)

# Pattern 2 — bare trailing number with no season keyword.
# Examples: "המתמחים 3", "הקינוח המושלם 1", "להיות איתה 3"
_TRAILING_NUMBER_RE = re.compile(r"^(.+?)\s+(\d+)\s*$", re.UNICODE)

# Hebrew column header → document field name mapping
COLUMN_MAP = {
    "מספר פרק": "episode_number",
    "מס' פרק":   "episode_number",   # short form used in some tabs
    "יום בשבוע": "day_of_week",
    "יום":       "day_of_week",       # short form
    "תאריך": "date",
    "בפרומו": "promo_text",
    "אודישנים": "promo_text",         # some tabs use section title as the promo column name
    "נקודת פתיחה": "opening_point",
    "רייטינג פרק": "rating",
    "תחרות": "competition",
}

# Strip "עונה N- " or "עונה N– " prefix from column header text before COLUMN_MAP lookup.
# Handles tabs like יצאת צדיק where the column is named "עונה 8- בפרומו" instead of "בפרומו".
_SEASON_PREFIX_RE = re.compile(r"^עונה\s+\d+\s*[-–]\s*")


def _normalize_header(text: str) -> str:
    """Normalise an Excel column header before COLUMN_MAP lookup."""
    return _SEASON_PREFIX_RE.sub("", text).strip()

# Azure AI Search accepts at most 1000 documents per batch upload call
BATCH_SIZE = 500


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def validate_env() -> None:
    """Raise if any required environment variable is missing."""
    required = {
        "AZURE_SEARCH_ENDPOINT": AZURE_SEARCH_ENDPOINT,
        "AZURE_SEARCH_KEY": AZURE_SEARCH_KEY,
        "AZURE_SEARCH_INDEX_NAME": AZURE_SEARCH_INDEX_NAME,
        "AZURE_STORAGE_CONNECTION_STRING": AZURE_STORAGE_CONNECTION_STRING,
        "AZURE_STORAGE_CONTAINER_NAME": AZURE_STORAGE_CONTAINER_NAME,
        "EXCEL_BLOB_NAME": EXCEL_BLOB_NAME,
        "AZURE_OPENAI_ENDPOINT": AZURE_OPENAI_ENDPOINT,
        "AZURE_OPENAI_KEY": AZURE_OPENAI_KEY,
        "AZURE_OPENAI_EMBEDDING_DEPLOYMENT": AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        raise EnvironmentError(
            f"Missing required environment variables: {', '.join(missing)}. "
            "Check your .env file."
        )


def download_blob(connection_string: str, container: str, blob_name: str) -> bytes:
    """Download a blob from Azure Blob Storage and return its raw bytes."""
    logger.info(f"Downloading '{blob_name}' from container '{container}' ...")
    blob_service = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service.get_blob_client(container=container, blob=blob_name)
    data = blob_client.download_blob().readall()
    logger.info(f"  Downloaded {len(data):,} bytes.")
    return data


def make_document_id(show_name: str, season: str, episode_number: str, row_index: int) -> str:
    """
    Generate a stable, unique document ID by hashing the identifying fields.
    row_index is included as a fallback to guarantee uniqueness when episode
    numbers are missing or duplicated.
    """
    raw = f"{show_name}|{season}|{episode_number}|{row_index}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:40]


def cell_value(cell: Any) -> str:
    """Return a cell's value as a stripped string, or empty string if None."""
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def embed_texts(
    http: httpx.Client,
    texts: list[str],
) -> list[list[float] | None]:
    """
    Embed a list of strings in batches via the Azure OpenAI REST API.

    Returns a list of float vectors aligned with the input list.
    Entries whose source text is empty (or whitespace-only) get None —
    the caller omits that field rather than storing null in the index.
    """
    vectors: list[list[float] | None] = [None] * len(texts)

    # Only embed non-empty texts; track their original positions
    non_empty = [(i, t) for i, t in enumerate(texts) if t.strip()]

    url = (
        f"{AZURE_OPENAI_ENDPOINT.rstrip('/')}/openai/deployments/"
        f"{AZURE_OPENAI_EMBEDDING_DEPLOYMENT}/embeddings"
        f"?api-version={AZURE_OPENAI_API_VERSION}"
    )

    for batch_start in range(0, len(non_empty), EMBED_BATCH_SIZE):
        batch = non_empty[batch_start : batch_start + EMBED_BATCH_SIZE]
        indices, batch_texts = zip(*batch)

        response = http.post(url, json={"input": list(batch_texts)})
        response.raise_for_status()
        data = response.json()["data"]

        # Azure OpenAI returns items sorted by their index in the request
        for item in data:
            vectors[indices[item["index"]]] = item["embedding"]

    return vectors


def parse_tab_name(tab_name: str) -> tuple[str, str]:
    """
    Split a tab title into (show_name, season).

    Tried in order:
      1. עונה/עונות keyword  → 'ארץ נהדרת עונה 18'      → ('ארץ נהדרת', '18')
                              → 'מאסטר שף עונה 9 VIP'    → ('מאסטר שף', '9 VIP')
                              → 'המטבח המנצח VIP - עונות 2 ו-3' → ('המטבח המנצח VIP', '2-3')
      2. Bare trailing number → 'המתמחים 3'               → ('המתמחים', '3')
      3. No number at all     → 'אור ראשון'               → ('אור ראשון', '')
    """
    tab = tab_name.strip()

    # --- Try explicit season keyword first.
    m = _SEASON_KEYWORD_RE.match(tab)
    if m:
        show_name = m.group(1).strip()
        season_raw = m.group(2).strip()
        # Normalise "2 ו-3" → "2-3"
        season = re.sub(r"\s*ו-\s*", "-", season_raw)
        return show_name, season

    # --- Fall back to a bare trailing number.
    m = _TRAILING_NUMBER_RE.match(tab)
    if m:
        return m.group(1).strip(), m.group(2).strip()

    # --- No season information found.
    return tab, ""


def parse_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    source_file: str,
) -> list[dict]:
    """
    Parse a single worksheet and return a list of document dicts.

    Sheet layout:
        Row 1  — merged title cell (e.g. "מעקב פרומו של התוכנית '…' עונה N")
        Row 2  — column headers (Hebrew names)
        Row 3+ — data rows
    """
    # --- Extract show name and season from the tab title.
    show_name, season = parse_tab_name(sheet.title)

    # --- Read column headers from row 2 to build a column-index → field-name map.
    header_map: dict[int, str] = {}  # 1-based column index → document field name
    for col_idx, cell in enumerate(sheet[2], start=1):
        header_text = _normalize_header(cell_value(cell))
        if header_text in COLUMN_MAP:
            header_map[col_idx] = COLUMN_MAP[header_text]

    if not header_map:
        logger.warning(
            f"  WARNING: No recognised column headers found in tab '{sheet.title}'. "
            "Skipping tab."
        )
        return []

    # --- Iterate data rows (row 3 onward) and build documents.
    documents: list[dict] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3), start=3):
        # Skip entirely empty rows
        row_values = [cell_value(c) for c in row]
        if not any(row_values):
            continue

        doc: dict[str, str] = {
            "show_name": show_name,
            "season": season,
            "source_file": source_file,
        }

        # Map each cell to its corresponding field using header_map
        for col_idx, cell in enumerate(row, start=1):
            field_name = header_map.get(col_idx)
            if field_name:
                doc[field_name] = cell_value(cell)

        # Ensure all expected fields are present (default to empty string)
        for field in COLUMN_MAP.values():
            doc.setdefault(field, "")

        # Generate a stable unique ID for this document
        doc["id"] = make_document_id(
            show_name,
            season,
            doc.get("episode_number", ""),
            row_idx,
        )

        documents.append(doc)

    return documents


def upload_in_batches(search_client: SearchClient, documents: list[dict]) -> int:
    """
    Upload documents to Azure AI Search in batches.
    Returns the total number of successfully uploaded documents.
    """
    total_uploaded = 0
    for start in range(0, len(documents), BATCH_SIZE):
        batch = documents[start : start + BATCH_SIZE]
        results = search_client.upload_documents(documents=batch)
        succeeded = sum(1 for r in results if r.succeeded)
        failed = len(batch) - succeeded
        total_uploaded += succeeded
        if failed:
            logger.warning(f"  WARNING: {failed} document(s) failed to upload in this batch.")
    return total_uploaded


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    validate_env()

    # Download the Excel file into memory — no temp file needed
    excel_bytes = download_blob(
        AZURE_STORAGE_CONNECTION_STRING,
        AZURE_STORAGE_CONTAINER_NAME,
        EXCEL_BLOB_NAME,
    )

    # Open the workbook from the in-memory bytes; data_only=True resolves formulas
    workbook = openpyxl.load_workbook(
        filename=io.BytesIO(excel_bytes), data_only=True
    )
    logger.info(f"Workbook loaded. Found {len(workbook.sheetnames)} sheet(s): {workbook.sheetnames}\n")

    # Build the Azure AI Search client once and reuse it for all tabs
    credential = AzureKeyCredential(AZURE_SEARCH_KEY)
    search_client = SearchClient(
        endpoint=AZURE_SEARCH_ENDPOINT,
        index_name=AZURE_SEARCH_INDEX_NAME,
        credential=credential,
    )

    # Build an HTTP client for the Azure OpenAI embeddings endpoint
    openai_client = httpx.Client(
        headers={"api-key": AZURE_OPENAI_KEY, "Content-Type": "application/json"},
        timeout=60.0,
    )

    grand_total = 0
    tabs_with_data = 0
    tabs_skipped = 0

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        logger.info(f"Processing tab: '{sheet_name}'")

        documents = parse_sheet(sheet, source_file=EXCEL_BLOB_NAME)

        if not documents:
            logger.info("  No documents to upload for this tab.\n")
            tabs_skipped += 1
            continue

        # Embed all promo_text values for this tab in one batched call
        promo_texts = [doc.get("promo_text", "") for doc in documents]
        vectors = embed_texts(openai_client, promo_texts)
        embedded = 0
        for doc, vector in zip(documents, vectors):
            if vector is not None:
                doc["promo_vector"] = vector
                embedded += 1

        tabs_with_data += 1
        logger.info(f"  Parsed {len(documents)} row(s), embedded {embedded}. Uploading ...")
        uploaded = upload_in_batches(search_client, documents)
        grand_total += uploaded
        logger.info(f"  Uploaded {uploaded}/{len(documents)} document(s).\n")

    total_tabs = len(workbook.sheetnames)
    logger.info("=" * 52)
    logger.info(f"  Total tabs in workbook : {total_tabs}")
    logger.info(f"  Tabs with data uploaded: {tabs_with_data}")
    logger.info(f"  Tabs skipped (no rows) : {tabs_skipped}")
    logger.info(f"  Total documents        : {grand_total}")
    logger.info("=" * 52)


if __name__ == "__main__":
    main()
